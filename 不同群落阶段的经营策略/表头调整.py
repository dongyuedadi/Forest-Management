import os
import shutil
import pandas as pd
from openpyxl import load_workbook


def needs_adjustment(file_path):
    """检查文件是否需要调整表头"""
    try:
        # 读取第一行作为表头
        df = pd.read_excel(file_path, nrows=1)
        current_headers = list(df.columns)
        target_headers = ['树号', '树种', '状态', '胸径', '树高', '死枝', '活枝', '东', '南', '西', '北', 'X', 'Y']

        # 检查当前表头是否与目标表头完全一致
        return current_headers != target_headers
    except Exception as e:
        print(f"检查文件 {file_path} 时出错: {str(e)}")
        return True  # 如果检查出错，默认需要处理


def adjust_excel_headers(input_file, output_file):
    """调整Excel表头格式"""
    # 使用openpyxl处理合并单元格的表头
    wb = load_workbook(input_file)
    ws = wb.active

    # 获取合并单元格信息
    merged_ranges = ws.merged_cells.ranges

    # 处理表头行（假设表头在第1行和第2行）
    header_rows = []
    for row in ws.iter_rows(min_row=1, max_row=2, values_only=True):
        header_rows.append(list(row))

    # 展平合并单元格的值
    for merge in merged_ranges:
        if merge.min_row == 1:  # 只处理第一行的合并单元格
            top_left_value = ws.cell(merge.min_row, merge.min_col).value
            for row in range(merge.min_row, merge.max_row + 1):
                for col in range(merge.min_col, merge.max_col + 1):
                    if row == merge.min_row and col == merge.min_col:
                        continue  # 保留左上角单元格的值
                    if row <= len(header_rows) and col <= len(header_rows[row - 1]):
                        header_rows[row - 1][col - 1] = top_left_value

    # 创建新的表头行
    new_header = []
    for col_idx in range(len(header_rows[0])):
        if header_rows[1][col_idx]:  # 如果有第二级表头
            col_name = f"{header_rows[0][col_idx]}_{header_rows[1][col_idx]}"
        else:
            col_name = header_rows[0][col_idx]
        new_header.append(col_name)

    # 读取数据（跳过表头行）
    data_rows = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        data_rows.append(list(row))

    # 创建DataFrame
    df = pd.DataFrame(data_rows, columns=new_header)

    # 定义列名映射（根据你的实际需求调整）
    column_mapping = {
        '树号_树号': '树号',
        '树种_树种': '树种',
        '状态_状态': '状态',
        '胸径_胸径': '胸径',
        '树高_树高': '树高',
        '枝下高（m）_死枝': '死枝',
        '枝下高（m）_活枝': '活枝',
        '冠幅（m）_东': '东',
        '冠幅（m）_南': '南',
        '冠幅（m）_西': '西',
        '冠幅（m）_北': '北',
        '坐标_X': 'X',
        '坐标_Y': 'Y'
    }

    # 重命名列
    df = df.rename(columns=column_mapping)

    # 选择需要的列
    final_columns = ['树号', '树种', '状态', '胸径', '树高', '死枝', '活枝', '东', '南', '西', '北', 'X', 'Y']
    df = df[[col for col in final_columns if col in df.columns]]

    # 保存到新文件
    df.to_excel(output_file, index=False)
    print(f"表头已调整并保存为 {output_file}")


def process_excel_files(input_folder, output_folder):
    """处理文件夹中的所有Excel文件"""
    # 创建输出文件夹（如果不存在）
    os.makedirs(output_folder, exist_ok=True)

    # 遍历输入文件夹中的所有Excel文件
    for filename in os.listdir(input_folder):
        if filename.endswith('.xlsx'):
            input_path = os.path.join(input_folder, filename)
            output_filename = f"调整后_{filename}"
            output_path = os.path.join(output_folder, output_filename)

            # 检查文件是否需要调整
            if not needs_adjustment(input_path):
                # 如果不需要调整，直接复制文件
                shutil.copy2(input_path, output_path)
                print(f"文件 {filename} 表头已符合要求，直接复制")
            else:
                # 如果需要调整，进行处理
                try:
                    adjust_excel_headers(input_path, output_path)
                except Exception as e:
                    print(f"处理文件 {filename} 时出错: {str(e)}")

    print("所有文件处理完成！")


# 设置文件夹路径
input_folder = r'C:\Users\hys5637428\Desktop\毕业论文\2022年帽儿山21块地原始数据-采伐设计-材积表-蓄积-采伐木选择\2022年帽儿山-21块地原始数据'
output_folder = r'C:\Users\hys5637428\Desktop\毕业论文\调整后数据'

# 处理文件
process_excel_files(input_folder, output_folder)